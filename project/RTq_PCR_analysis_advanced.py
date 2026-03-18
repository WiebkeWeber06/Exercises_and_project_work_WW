#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Oct  9 14:28:34 2024

@author: wweber
"""

import pandas as pd
import numpy as np
import os
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
import re
from datetime import datetime
from openpyxl.drawing.image import Image
from collections import defaultdict
from itertools import cycle
from decimal import Decimal
from collections import Counter
from pathlib import Path
import matplotlib
import matplotlib.ticker as ticker
from matplotlib.lines import Line2D
import glob
from scipy import stats
import statsmodels.api as sm
from statsmodels.formula.api import ols
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from scipy.stats import kruskal
from scipy.stats import shapiro, levene
import scikit_posthocs as sp
from statsmodels.stats.multitest import multipletests
from itertools import combinations
import matplotlib.patches as patches
import matplotlib.colors as mcolors
from scipy.stats import ttest_ind, f_oneway
import statistics

#------------------------------------------------------------------------------
###functions
def extract_date_from_path(file_path):
    # Use regex to extract the date from the file path
    date_match = re.search(r'(\d{2}-\d{2}-\d{2})', file_path)
    
    # Return the date if found, else return None
    return date_match.group(1) if date_match else None

###collect sample names without technical replicate names
def collect_values_based_on_name(df):
    outer_dict = {}
    for index, row in df.iterrows():
        sample_name = row[0]
        # Exit the loop if the sample name is empty or NaN
        if pd.isna(sample_name) or str(sample_name).strip() == "":
           print(f"Stopping loop at row {index} due to empty sample name.")
           break
        sample_name = str(sample_name).strip()
        # Exclude base names that start with "Sample" to exclude empty rows
        if sample_name.startswith("Sample"):
            continue
        base_name = sample_name[:-2]
        if base_name not in outer_dict:
            outer_dict[base_name] = {}   
          
    return outer_dict 

###Add CT-values for all technical replicates
def add_inner_dicts(df, outer_dict):
    for index, row in df.iterrows():
            sample_name = str(row[0]).strip()
            gene_name = str(row[1]).strip()
            CT_value = row[2]
            if isinstance(CT_value, str):
                CT_value = CT_value.strip()
            for key, value in outer_dict.items():
                if key == sample_name[:-2]:
                    if gene_name not in outer_dict[key]:
                        outer_dict[key][gene_name] = []
                    outer_dict[key][gene_name].append(CT_value)

    return outer_dict            

###calculate mean for technical replicates        
def calculate_mean_ct_values(outer_dict):
    mean_dict = {}    
    for base_name, gene_dict in outer_dict.items():
        mean_dict[base_name] = {}
        for gene_name, ct_values in gene_dict.items():
            ct_values_filtered = [value for value in ct_values if value != '-']
            if len(ct_values_filtered) > 0:
                ct_values_float = [float(value) for value in ct_values_filtered]
                mean_ct_value = statistics.mean(ct_values_float)
                mean_dict[base_name][gene_name] = mean_ct_value
            else:
                mean_dict[base_name][gene_name] = None
    return mean_dict

###Calculate relative expression for technical replicates
def relative_expression_strength(means_dict):
    delta_ct_dict = {}
    for base_name, gene_dict in means_dict.items():
        if 'RPL13' in gene_dict and gene_dict['RPL13'] is not None:
            
            value_18S = gene_dict['RPL13']
            if value_18S is not None:
               delta_ct_dict[base_name] = {} 
            
            for gene_name, mean_value in gene_dict.items():
                if gene_name != 'RPL13' and mean_value is not None:
                    delta_ct = mean_value - value_18S
                    transformed_value = 2 ** (-delta_ct)
                    delta_ct_dict[base_name][gene_name] = transformed_value
                    
    return delta_ct_dict

def calculate_relative_expression_with_error(outer_dict):
    relative_expression = {}
    error_bars = {}

    for base_name, gene_dict in outer_dict.items():
        relative_expression[base_name] = {}
        error_bars[base_name] = {}

        if 'RPL13' not in gene_dict:
            continue

        ref_cts = [float(ct) for ct in gene_dict['RPL13'] if ct != '-']
        if not ref_cts:
            continue

        for gene_name, ct_list in gene_dict.items():
            if gene_name == 'RPL13':
                continue

            valid_cts = [float(ct) for ct in ct_list if ct != '-']
            if not valid_cts or len(valid_cts) != len(ref_cts):
                continue  # skip if lengths mismatch (can also pad to match lengths if needed)

            delta_cts = [gene_ct - ref_ct for gene_ct, ref_ct in zip(valid_cts, ref_cts)]
            rel_exprs = [2 ** (-delta_ct) for delta_ct in delta_cts]

            mean_expr = np.mean(rel_exprs)
            std_expr = np.std(rel_exprs, ddof=1)

            relative_expression[base_name][gene_name] = mean_expr
            error_bars[base_name][gene_name] = std_expr

    return relative_expression, error_bars

def calculate_change_vs_baseline(relative_expression, compare_basename, error_bars_rE):
    fold_change = {}
    error_change = {}

    for base_name, genes in relative_expression.items():
        fold_change[base_name] = {}
        error_change[base_name] = {}

        for gene_name, value in genes.items():
            # Skip if baseline does not have this gene
            if gene_name not in relative_expression[compare_basename]:
                continue

            # Compute fold change
            baseline_value = relative_expression[compare_basename][gene_name]
            if baseline_value == 0:
                continue  # Avoid division by zero

            fold = value / baseline_value
            fold_change[base_name][gene_name] = fold

            # Get the error bars (standard deviations) from the error_bars_rE dictionary
            error_sample = error_bars_rE.get(base_name, {}).get(gene_name, 0)
            error_baseline = error_bars_rE.get(compare_basename, {}).get(gene_name, 0)

            # Propagate error (standard deviation) using the error propagation formula
            propagated_error = fold * np.sqrt((error_sample / value)**2 + (error_baseline / baseline_value)**2)
            error_change[base_name][gene_name] = propagated_error

    return fold_change, error_change


def relative_expression_per_replicate(replicates_dict):
    re_dict = {}

    for base_name, gene_dict in replicates_dict.items():
        if 'RPL13' in gene_dict and gene_dict['RPL13'] is not None:
            # Convert RPL13 values to floats
            try:
                reference_vals = [float(x) for x in gene_dict['RPL13']]
            except (ValueError, TypeError):
                continue  # Skip this strain if values can't be converted

            num_reps = len(reference_vals)
            re_dict[base_name] = {}

            for gene_name, gene_vals in gene_dict.items():
                if gene_name != 'RPL13' and gene_vals is not None:
                    try:
                        target_vals = [float(x) for x in gene_vals]
                    except (ValueError, TypeError):
                        continue  # Skip if conversion fails

                    # Check matching lengths
                    if len(target_vals) != num_reps:
                        continue

                    # Calculate rE per replicate
                    re_list = []
                    for ref, target in zip(reference_vals, target_vals):
                        delta_ct = target - ref
                        transformed = 2 ** (-delta_ct)
                        re_list.append(transformed)

                    re_dict[base_name][gene_name] = re_list

    return re_dict

###Check for biological replicates
def check_for_biological_replicates(outer_dict):
    replicate_count = {}
    for basename in outer_dict.keys():
        match = re.match(r'^([A-Za-z]+)\d+_(.*)', basename)
        if match:
            condition = match.group(2)
            if condition in replicate_count:
                replicate_count[condition] += 1
            else:
                replicate_count[condition] = 1
    return any(count > 1 for count in replicate_count.values())                

      

###Calculate relative expression for biological replicates
def relative_expression_mean_for_biological_replicates(delta_ct_dict):
    extracted_basenames = {}
    for basename, parts_dict in relative_expression.items():
        parts = basename.split('_', 1)
        extracted_basenames[basename] = {}
        if len(parts) == 2:
            extracted_basenames[basename]['part1'] = parts[0][:-2]
            extracted_basenames[basename]['part2'] = parts[1][:-2]
    
    grouped_values = {}
    for basename, parts in extracted_basenames.items():
        part1 = parts['part1']
        part2 = parts['part2']
        key = (part1, part2)
    
        if key not in grouped_values:
            grouped_values[key] = {
            'sum': 0.0,
                'count': 0
                }
        
        if basename in relative_expression:
            relative_value = float(relative_expression[basename]['LHCBM9'])
            grouped_values[key]['sum'] += relative_value
            grouped_values[key]['count'] += 1
        
    mean_relative_expression = {}
    for key, values in grouped_values.items():
        part1, part2 = key
        mean_value = values['sum'] / values['count']
        mean_relative_expression[f"{part1}_{part2}"] = mean_value
       
    return mean_relative_expression     


###Calculate change in expression according to a set comparison value
def change_in_expression(relative_expression, compare_basename, compare_gene_name):
   change_in_expression = {}
   compare_key_dict = relative_expression.get(compare_basename)
   compare_key_value = compare_key_dict.get(compare_gene_name)
   if compare_key_value is None:
       raise ValueError(f"Base name '{compare_basename}' not found in relative_expression.")
   for base_name, gene_dict in relative_expression.items():        
       change_in_expression[base_name] = {}
       for gene_name, expression_value in gene_dict.items():
           change_in_expression_value = expression_value / compare_key_value
           change_in_expression[base_name][gene_name] = change_in_expression_value
   return change_in_expression        
            
def auto_adjust_column_widths(excel_file_path):
    """Auto-adjusts column widths for all sheets in the given Excel file based on content length."""
    wb = load_workbook(excel_file_path)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        # Adjust width of index column (assumed to be column A)
        max_length_index = max((len(str(cell.value)) for cell in ws['A'] if cell.value), default=0)
        ws.column_dimensions['A'].width = max_length_index + 2

        # Adjust width for the rest of the columns
        for col_cells in ws.iter_cols(min_col=2):  # Skip index column
            max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=0)
            column_letter = col_cells[0].column_letter
            ws.column_dimensions[column_letter].width = max_length + 2

    # Save changes
    wb.save(excel_file_path)
    
    
    
### Stats
# --- Function 1: t-tests (timepoints within strain) ---
def ttest_across_timepoints_within_strain(df_re, df_result_dict_CT):
    meta_tp = []
    pvals_tp = []

    for gene in df_re["Gene"].unique():
        gene_group = df_re[df_re["Gene"] == gene]

        for strain in gene_group["Strain"].unique():
            strain_group = gene_group[gene_group["Strain"] == strain]
            timepoints = sorted(strain_group["Timepoint"].unique())

            for tp1, tp2 in combinations(timepoints, 2):
                data1 = strain_group[strain_group["Timepoint"] == tp1]["Relative_Expression"].dropna()
                data2 = strain_group[strain_group["Timepoint"] == tp2]["Relative_Expression"].dropna()

                if len(data1) >= 2 and len(data2) >= 2:
                    stat, pval = ttest_ind(data1, data2, equal_var=False)
                    pvals_tp.append(pval)
                    meta_tp.append((gene, strain, tp1, tp2))

    reject_tp, pvals_corr_tp, _, _ = multipletests(pvals_tp, alpha=0.05, method='bonferroni')

    df_ttest = pd.DataFrame([
        {
            "Gene": gene,
            "Strain": strain,
            "Timepoint1": tp1,
            "Timepoint2": tp2,
            "Raw p-value": raw,
            "Bonferroni p-value": corr,
            "Significant": sig
        }
        for (gene, strain, tp1, tp2), raw, corr, sig in zip(meta_tp, pvals_tp, pvals_corr_tp, reject_tp)
    ])

    return df_ttest

# --- Function 2: ANOVA + Tukey (strains within timepoint) ---
def anova_and_tukey_across_strains_within_timepoint(df_re):
    anova_results = []
    tukey_results = []

    for gene in df_re["Gene"].unique():
        gene_group = df_re[df_re["Gene"] == gene]

        for tp in gene_group["Timepoint"].unique():
            tp_group = gene_group[gene_group["Timepoint"] == tp]
            strains = tp_group["Strain"].unique()

            group_data = [tp_group[tp_group["Strain"] == s]["Relative_Expression"].dropna().values for s in strains]

            # Only proceed if all groups have at least 2 replicates
            if all(len(g) >= 2 for g in group_data) and len(group_data) > 1:
                # ANOVA
                anova_stat, anova_pval = f_oneway(*group_data)
                anova_results.append({
                    "Gene": gene,
                    "Timepoint": tp,
                    "ANOVA F": anova_stat,
                    "ANOVA p-value": anova_pval
                })

                if anova_pval < 0.05:
                    # Tukey HSD
                    tukey = pairwise_tukeyhsd(
                        endog=tp_group["Relative_Expression"],
                        groups=tp_group["Strain"],
                        alpha=0.05
                    )

                    for res in tukey.summary().data[1:]:  # skip header
                        tukey_results.append({
                            "Gene": gene,
                            "Timepoint": tp,
                            "Strain1": res[0],
                            "Strain2": res[1],
                            "Mean Diff": res[2],
                            "p-adj": res[3],
                            "Lower": res[4],
                            "Upper": res[5],
                            "Reject H0": res[6]
                        })

    df_anova = pd.DataFrame(anova_results)
    df_tukey = pd.DataFrame(tukey_results)

    return df_anova, df_tukey

# Track how many annotation lines are already on each bar
y_offset_tracker = {}
def get_annotation_height(t, g1, g2, base_y):
    # Get how many annotations already used for each bar
    k1 = (t, g1)
    k2 = (t, g2)
    
    # Get current max offset used for either group
    current_offset = max(y_offset_tracker.get(k1, 0), y_offset_tracker.get(k2, 0))
    
    # Compute height
    y = base_y + current_offset * 0.7  # You can tweak the 0.2 gap
    
    # Update tracker
    y_offset_tracker[k1] = current_offset + 1
    y_offset_tracker[k2] = current_offset + 1
    
    return y
        
#------------------------------------------------------------------------------
###Get input data

#print('Please enter an RTq PCR raw data file for analysis: ')
#input_file_path = input()
file_path = '/prj/lhcbm9/02_Master_project_08-2024_03-2025/04_RTq_PCR/00_OEx/results_20-11-24.xlsx'
output_file_path = '/prj/lhcbm9/02_Master_project_08-2024_03-2025/04_RTq_PCR/00_OEx/'
compare_basename = 'CC-1883_t0_1'

date_file_name = extract_date_from_path(file_path)

#print('Please enter a comparison sample (negative control) for your analysis')'

#print('Please enter a comparison sample (negative control) for your analysis')
#compare_basename = input()

#------------------------------------------------------------------------------
###Read dataframe

df = pd.read_excel(file_path, sheet_name='Tabelle1', usecols=[2, 3, 4])
df.columns = df.columns.str.strip()

df['Gene Name'] = df['Gene Name']
df['Sample Name'] = df['Sample Name']
df = df[df['Gene Name'].str.strip() != 'LHCBM1']
df = df[df['Sample Name'].str.contains('CC-1883|water', case=False, na=False)]
df = df.sort_values(by=["Sample Name", "Gene Name"])


#------------------------------------------------------------------------------
###code
#---------------------------------------------------------------------------------------------------
###collect sample names without technical replicate names
result_dict = collect_values_based_on_name(df)

###Add CT-values for all technical replicates
result_dict_CT = add_inner_dicts(df, result_dict)
result_dict_CT_df = pd.DataFrame(result_dict_CT)


#---------------------------------------------------------------------------------------------------
### releative expression

# calculate mean CT values for each gene for technical replicates   
means_CT_values = calculate_mean_ct_values(result_dict_CT)

# Calculate relative expression means per gene and error bars
relative_expression, error_bars_rE = calculate_relative_expression_with_error(result_dict_CT)

### Make dataframe
# Combine relative expression and error bars into a tidy dataframe
combined_data = []

for basename in relative_expression:
    for gene in relative_expression[basename]:
        mean_expr = relative_expression[basename][gene]
        error = error_bars_rE.get(basename, {}).get(gene, None)

        combined_data.append({
            "Basename": basename,
            "Gene": gene,
            "Relative_Expression": mean_expr,
            "Error_rE": error
        })

# Create the DataFrame
rE_all_means_and_errors_df = pd.DataFrame(combined_data)

# Extract metadata (Strain, Timepoint, Replicate) from Basename
rE_all_means_and_errors_df[["Strain", "Timepoint", "Replicate"]] = rE_all_means_and_errors_df["Basename"].str.split("_", expand=True)


#--------------------------------------------------------------
### Calculate relative expression for technical replicates each
rE_replicates = relative_expression_per_replicate(result_dict_CT)

# make rE_replicates a dataframe
rows = []

for key, genes in rE_replicates.items():
    try:
        strain, timepoint, replicate = key.split("_")
    except ValueError:
        print(f"Skipping invalid key: {key}")
        continue

    
    for gene, expr_list in genes.items():
        for i, val in enumerate(expr_list):
            rows.append({
                "Strain": strain,
                "Timepoint": timepoint,
                "Replicate": i + 1,
                "Gene": gene,
                "Relative_Expression": val
            })

df_rE_replicates = pd.DataFrame(rows)


#---------------------------------------------------------------------------------------------------
### Fold Change

# Calculate fold change means per gene and error bars
fold_change, error_bars_fC = calculate_change_vs_baseline(relative_expression, compare_basename, error_bars_rE)

# Create an empty list to store the combined fold change and error data
combined_fold_change_data = []

# Loop over the fold change data and combine it with the corresponding errors
for basename in fold_change:
    for gene in fold_change[basename]:
        fold_value = fold_change[basename][gene]
        error = error_bars_fC.get(basename, {}).get(gene, None)

        combined_fold_change_data.append({
            "Basename": basename,
            "Gene": gene,
            "Fold_Change": fold_value,
            "Error_FC": error
        })

# Create the DataFrame for fold change data
fold_change_all_means_and_errors_df = pd.DataFrame(combined_fold_change_data)

# Extract metadata (Strain, Timepoint, Replicate) from Basename
fold_change_all_means_and_errors_df[["Strain", "Timepoint", "Replicate"]] = fold_change_all_means_and_errors_df["Basename"].str.split("_", expand=True) 


#---------------------------------------------------------------------------------------------------
##output in excel
output_file_path_full = os.path.join(output_file_path, f"{output_file_path}_RTq_PCR_{date_file_name}_output.xlsx")    
    

with pd.ExcelWriter(output_file_path_full, engine='openpyxl', mode='w') as writer:  
    result_dict_CT_df.to_excel(writer, sheet_name='CT_values', startrow=1, startcol=1, index=True)    
    rE_all_means_and_errors_df.to_excel(writer, sheet_name='rE', startrow=1, startcol=1, index=True)    
    fold_change_all_means_and_errors_df.to_excel(writer, sheet_name='fold_change', startrow=1, startcol=1, index=True)       
    
    
# Adjust the column widths
auto_adjust_column_widths(output_file_path_full)  

#---------------------------------------------------------------------------------------------------
### Prepare Fold Change Data

# Assuming df_rE_replicates is your dataframe with replicate-level data

# Step 1: Filter for baseline expression (T0 of strain 'CC-1883')
baseline_df = df_rE_replicates[(df_rE_replicates["Strain"] == "CC-1883") & (df_rE_replicates["Timepoint"] == "t0")]

# Step 2: Merge baseline expression with the original dataframe
baseline_df = baseline_df[["Gene", "Relative_Expression", "Replicate"]]  # Keep relevant columns
baseline_df = baseline_df.rename(columns={"Relative_Expression": "Baseline_Expression"})

# Merge baseline expression into the original dataframe
df_with_baseline = df_rE_replicates.merge(
    baseline_df,
    on=["Gene", "Replicate"],
    how="left"
)

# Step 3: Calculate log2 Fold Change for each replicate
df_with_baseline["Log2_FC"] = np.log2(df_with_baseline["Relative_Expression"] / df_with_baseline["Baseline_Expression"])



# Stats
#---------------------------------------------------------------------------------------------------
df_ttests = ttest_across_timepoints_within_strain(df_rE_replicates, result_dict_CT_df)
df_anova, df_tukey = anova_and_tukey_across_strains_within_timepoint(rE_all_means_and_errors_df)

# Stats relative Expression
#---------------------------------------------------------------------------------------------------
# Store ANOVA results
anova_results_rE = {}
tukey_results_rE = {}


    
if len(df_rE_replicates['Timepoint'].unique()) > 1: 
    # Fit the model: change 'Group' to whatever your actual grouping column is
    model = ols('Relative_Expression ~ C(Timepoint)', data=df_rE_replicates).fit()
    anova_table = sm.stats.anova_lm(model, typ=2)
    
    anova_results_rE = anova_table
    
    # Check if the ANOVA is significant
    pval = anova_table['PR(>F)'][0]
    if pval < 0.05:
        # Run Tukey HSD post-hoc test
        tukey = pairwise_tukeyhsd(endog=df_rE_replicates['Relative_Expression'],
                                  groups=df_rE_replicates['Timepoint'],
                                  alpha=0.05)
        tukey_df = pd.DataFrame(data=tukey.summary().data[1:], columns=tukey.summary().data[0])
        tukey_results_rE = tukey_df
            
    if len(anova_results_rE) > 0:       
        # Combine ANOVA tables into a single DataFrame with Timepoint as a column
        anova_results_rE_df = pd.concat(
            {k: v for k, v in anova_results_rE.items()},
            names=['Timepoint', 'Source']
        ).reset_index()
        
        
        tukey_results_rE_df = pd.concat(
            {k: v for k, v in tukey_results_rE.items()},
            names=['Timepoint']
        ).reset_index()







#---------------------------------------------------------------------------------------------------
### colors

# OEx codon opt L9
palette_dict = {'20-11-24':'#B200FF',
                '07-01-25': '#FF1493'}

# natural L9
palette_dict_2 = {'20-11-24':'#00BFB6',
                '07-01-25': '#00FFCE'}

# Updated, high-contrast, distinguishable colors
custom_palette_basenames = [
    "#FFBB00",  # yellow CC-1883
    "#80ffdb", "#07BEB8",  # turquoise KOs
    "#80CBFF", "#0A84E6", "#003A7E"   # blue shades OEx
    
]


#---------------------------------------------------------------------------------------------------
# Plotting clustered barplot

x_pos = []

plt.figure(figsize=(10/2, 7))

plt.ylim(0, 25)

ax = sns.barplot(
    data=df_rE_replicates,
    x="Timepoint", y='Relative_Expression', hue="Strain",
    palette= custom_palette_basenames,
    capsize=0.3, errorbar='se',
    err_kws={'color': 'black',
             'linewidth' : 1})

# Now add error bars on top of actual bars
for bar, (_, row) in zip(ax.patches, fold_change_all_means_and_errors_df.iterrows()):
    x = bar.get_x() + bar.get_width() / 2
    x_pos.append(x)
    y = bar.get_height()
    err = row["Error_FC"]

    #ax.errorbar(
       # x, y, yerr=err,
       # fmt='none', capsize=10, color='black', linewidth=1
    #)
    
    #ax.errorbar(x, y, yerr=err, fmt='none', capsize=10, color='black', linewidth=1)

    # Pick custom offset based on the 'Date' column
    date_offset = row["Timepoint"]
    #offset = 0.0015 if row["Timepoint"] == "t0" else 0.0055
    
    plt.text(
        x,  # X-coordinate: center of the bar
        y + 1.5,                            # Y-coordinate: top of the bar
        f"{y:,.2f}",                # Format the value with commas
        ha='center', va='bottom',          # Align text to be centered
        fontsize=18, color='black'
    )
    
plt.legend(title="Strain",
    fontsize=18,
    loc='upper right',
    frameon = True,
    fancybox = True,
    #bbox_to_anchor=(0.5, -0.6),
    #columnspacing = 10,
    ncol=1, 
    title_fontsize=20,
    )

x_pos = sorted(x_pos)

x_pos_strain = []

# Iterate through the list in steps of 2
for i in range(0, len(x_pos), 2):
    # Average the current value and the next one (if it exists)
    avg = (x_pos[i] + x_pos[i + 1]) / 2 if i + 1 < len(x_pos) else x_pos[i]
    x_pos_strain.append(avg)
    

# Add a vertical dotted line after the fourth bar (index 3, between 3 and 4)
#plt.axvline(x=3.5, color='#333333', linestyle='dotted', linewidth=2)
#plt.axvline(x=8.5, color='#333333', linestyle='dotted', linewidth=2)

# Get mean heights to match with bar heights (Seaborn uses the mean for barplot)
group_means = df_rE_replicates.groupby(['Timepoint', 'Strain'])['Relative_Expression'].mean().reset_index()

# Map from (Timepoint, Strain) to x-bar center
x_pos_map = {}

for bar in ax.patches:
    height = bar.get_height()
    # x position of the center of the bar
    x = bar.get_x() + bar.get_width() / 2

    # Try to find the (Timepoint, Strain) pair with this height
    # Allow a small tolerance because of float precision
    match = group_means[np.isclose(group_means['Relative_Expression'], height, atol=1e-4)]

    # If there's only one match, assign it
    if len(match) == 1:
        t = match['Timepoint'].values[0]
        s = match['Strain'].values[0]
        x_pos_map[(t, s)] = x
    

sig_comparisons = tukey_results_rE[tukey_results_rE['reject'] == True]

for i, row in sig_comparisons.iterrows():
    timepoint1 = row['group1']
    timepoint2 = row['group2']
    x1 = x_pos_map.get((timepoint1, 'CC-1883'))
    x2 = x_pos_map.get((timepoint2, 'CC-1883'))
    
    if x1 is None or x2 is None:
        continue  # Skip if mapping failed

    # Determine height for annotation line
    y1 = df_rE_replicates[(df_rE_replicates['Timepoint'] == timepoint1) &
                          (df_rE_replicates['Strain'] == 'CC-1883')]['Relative_Expression'].max()
    y2 = df_rE_replicates[(df_rE_replicates['Timepoint'] == timepoint2) &
                          (df_rE_replicates['Strain'] == 'CC-1883')]['Relative_Expression'].max()
    base_y = max(y1, y2) + 1.5
    y = get_annotation_height('CC-1883', timepoint1, timepoint2, base_y)
    h = 1
 
    ax.plot([x1, x1, x2, x2], [y, y + h, y + h, y], lw=1.5, c='black')
    ax.text((x1 + x2)/2, y + h + 0.05, '*', ha='center', va='bottom', fontsize=20)
     
    
    
    

plt.xlabel('Timepoint', fontsize=24, labelpad=20)
plt.ylabel('rE (compared to RPL13)', fontsize=24, labelpad=20)

plt.xticks(rotation = 0, fontsize=18)
plt.yticks(fontsize=18)

# Save the plot
plt.tight_layout()
output_file_path_pic = f"{output_file_path}RTq_PCR_analysis_plot_CC1883_LHCBM9.png"   
plt.savefig(output_file_path_pic, dpi=300, format='png')  # Save as PNG with high resolution

# Show the plot
plt.show()    

