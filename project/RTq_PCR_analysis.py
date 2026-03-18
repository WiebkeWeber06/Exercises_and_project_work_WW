#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Oct  9 14:28:34 2024

@author: wweber
"""
from collections import abc
from collections import defaultdict
import pandas as pd
import statistics
import numpy as np
import scipy
import scipy.sparse
#from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import re
from openpyxl import Workbook
from matplotlib.ticker import ScalarFormatter
from matplotlib.patches import Patch

#------------------------------------------------------------------------------
###functions

###collect sample names without technical replicate names
def collect_values_based_on_name(df):
    outer_dict = {}
    for index, row in df.iterrows():
        sample_name = row[2]
        # Exit the loop if the sample name is empty or NaN
        if pd.isna(sample_name) or str(sample_name).strip() == "":
           print(f"Stopping loop at row {index} due to empty sample name.")
           break
        sample_name = str(sample_name).strip()
        # Exclude base names that start with "Sample" to exclude empty rows
        if sample_name.startswith("Sample"):
            continue
        base_name = sample_name[:-2]
        if base_name not in outer_dict:
            outer_dict[base_name] = {}   
          
    return outer_dict 

###Add CT-values for all technical replicates
def add_inner_dicts(df, outer_dict):
    for index, row in df.iterrows():
            sample_name = str(row[2]).strip()
            gene_name = str(row[3]).strip()
            CT_value = row[5]
            if isinstance(CT_value, str):
                CT_value = CT_value.strip()
            for key, value in outer_dict.items():
                if key == sample_name[:-2]:
                    if gene_name not in outer_dict[key]:
                        outer_dict[key][gene_name] = []
                    outer_dict[key][gene_name].append(CT_value)

    return outer_dict            

###calculate mean for technical replicates        
def calculate_mean_ct_values(outer_dict):
    mean_dict = {}    
    for base_name, gene_dict in outer_dict.items():
        mean_dict[base_name] = {}
        for gene_name, ct_values in gene_dict.items():
            ct_values_filtered = [value for value in ct_values if value != '-']
            if len(ct_values_filtered) > 0:
                ct_values_float = [float(value) for value in ct_values_filtered]
                mean_ct_value = statistics.mean(ct_values_float)
                mean_dict[base_name][gene_name] = mean_ct_value
            else:
                mean_dict[base_name][gene_name] = None
    return mean_dict

###Calculate relative expression for technical replicates
def relative_expression_strength(means_dict):
    delta_ct_dict = {}
    for base_name, gene_dict in means_dict.items():
        if 'RPL13' in gene_dict and gene_dict['RPL13'] is not None:
            
            value_18S = gene_dict['RPL13']
            if value_18S is not None:
               delta_ct_dict[base_name] = {} 
            
            for gene_name, mean_value in gene_dict.items():
                if gene_name != 'RPL13' and mean_value is not None:
                    delta_ct = mean_value - value_18S
                    transformed_value = 2 ** (-delta_ct)
                    delta_ct_dict[base_name][gene_name] = transformed_value
                    
    return delta_ct_dict

def calculate_relative_expression_with_error(outer_dict):
    relative_expression = {}
    error_bars = {}

    for base_name, gene_dict in outer_dict.items():
        relative_expression[base_name] = {}
        error_bars[base_name] = {}

        if 'RPL13' not in gene_dict:
            continue

        ref_cts = [float(ct) for ct in gene_dict['RPL13'] if ct != '-']
        if not ref_cts:
            continue

        for gene_name, ct_list in gene_dict.items():
            if gene_name == 'RPL13':
                continue

            valid_cts = [float(ct) for ct in ct_list if ct != '-']
            if not valid_cts or len(valid_cts) != len(ref_cts):
                continue  # skip if lengths mismatch (can also pad to match lengths if needed)

            delta_cts = [gene_ct - ref_ct for gene_ct, ref_ct in zip(valid_cts, ref_cts)]
            rel_exprs = [2 ** (-delta_ct) for delta_ct in delta_cts]

            mean_expr = np.mean(rel_exprs)
            std_expr = np.std(rel_exprs, ddof=1)

            relative_expression[base_name][gene_name] = mean_expr
            error_bars[base_name][gene_name] = std_expr

    return relative_expression, error_bars

def relative_expression_per_replicate(replicates_dict):
    re_dict = {}

    for base_name, gene_dict in replicates_dict.items():
        if 'RPL13' in gene_dict and gene_dict['RPL13'] is not None:
            # Convert RPL13 values to floats
            try:
                reference_vals = [float(x) for x in gene_dict['RPL13']]
            except (ValueError, TypeError):
                continue  # Skip this strain if values can't be converted

            num_reps = len(reference_vals)
            re_dict[base_name] = {}

            for gene_name, gene_vals in gene_dict.items():
                if gene_name != 'RPL13' and gene_vals is not None:
                    try:
                        target_vals = [float(x) for x in gene_vals]
                    except (ValueError, TypeError):
                        continue  # Skip if conversion fails

                    # Check matching lengths
                    if len(target_vals) != num_reps:
                        continue

                    # Calculate rE per replicate
                    re_list = []
                    for ref, target in zip(reference_vals, target_vals):
                        delta_ct = target - ref
                        transformed = 2 ** (-delta_ct)
                        re_list.append(transformed)

                    re_dict[base_name][gene_name] = re_list

    return re_dict

###Check for biological replicates
def check_for_biological_replicates(outer_dict):
    replicate_count = {}
    for basename in outer_dict.keys():
        match = re.match(r'^([A-Za-z]+)\d+_(.*)', basename)
        if match:
            condition = match.group(2)
            if condition in replicate_count:
                replicate_count[condition] += 1
            else:
                replicate_count[condition] = 1
    return any(count > 1 for count in replicate_count.values())                

      

###Calculate relative expression for biological replicates
def relative_expression_mean_for_biological_replicates(delta_ct_dict):
    extracted_basenames = {}
    for basename, parts_dict in relative_expression.items():
        parts = basename.split('_', 1)
        extracted_basenames[basename] = {}
        if len(parts) == 2:
            extracted_basenames[basename]['part1'] = parts[0][:-2]
            extracted_basenames[basename]['part2'] = parts[1][:-2]
    
    grouped_values = {}
    for basename, parts in extracted_basenames.items():
        part1 = parts['part1']
        part2 = parts['part2']
        key = (part1, part2)
    
        if key not in grouped_values:
            grouped_values[key] = {
            'sum': 0.0,
                'count': 0
                }
        
        if basename in relative_expression:
            relative_value = float(relative_expression[basename]['LHCBM9'])
            grouped_values[key]['sum'] += relative_value
            grouped_values[key]['count'] += 1
        
    mean_relative_expression = {}
    for key, values in grouped_values.items():
        part1, part2 = key
        mean_value = values['sum'] / values['count']
        mean_relative_expression[f"{part1}_{part2}"] = mean_value
       
    return mean_relative_expression     


###Calculate change in expression according to a set comparison value
def change_in_expression(relative_expression, compare_basename, compare_gene_name):
   change_in_expression = {}
   compare_key_dict = relative_expression.get(compare_basename)
   compare_key_value = compare_key_dict.get(compare_gene_name)
   if compare_key_value is None:
       raise ValueError(f"Base name '{compare_basename}' not found in relative_expression.")
   for base_name, gene_dict in relative_expression.items():        
       change_in_expression[base_name] = {}
       for gene_name, expression_value in gene_dict.items():
           change_in_expression_value = expression_value / compare_key_value
           change_in_expression[base_name][gene_name] = change_in_expression_value
   return change_in_expression        
            





        
#------------------------------------------------------------------------------
###Get input data

#print('Please enter an RTq PCR raw data file for analysis: ')
#input_file_path = input()
input_file_path = '/prj/lhcbm9/02_Master_project_08-2024_03-2025/04_RTq_PCR/00_OEx/results_20-11-24.xlsx'
output_file_path = '/prj/lhcbm9/02_Master_project_08-2024_03-2025/04_RTq_PCR/00_OEx/'

#print('Please enter a comparison sample (negative control) for your analysis')'

#print('Please enter a comparison sample (negative control) for your analysis')
#compare_basename = input()

#------------------------------------------------------------------------------
###Read dataframe

df = pd.read_excel(input_file_path, sheet_name='Tabelle1')

#------------------------------------------------------------------------------
###
# Strip whitespace from column names
df.columns = df.columns.str.strip()

# Rename strains
rename_dict = {'pML2_603_' : 'YFP+L9-OEx-',
               'pML2_604_' : 'L9-OEx-',
               'CC-1883' : 'PS',
    }

for old, new in rename_dict.items():
    df['Sample Name'] = df['Sample Name'].str.replace(old, new, regex=True)

###collect sample names without technical replicate names
result_dict = collect_values_based_on_name(df)

###Add CT-values for all technical replicates
result_dict_CT = add_inner_dicts(df, result_dict)

###calculate mean for technical replicates   
means = calculate_mean_ct_values(result_dict_CT)

###Calculate relative expression for technical replicates
#relative_expression = relative_expression_strength(means)
relative_expression, error_bars = calculate_relative_expression_with_error(result_dict_CT)
df_error_bars = pd.DataFrame(error_bars)

###Check for biological replicates
if check_for_biological_replicates(result_dict_CT):
    #store relative_expression dictionary in a new variable to clean the
    #varaible for further calculation to simplify the code
    relative_expression_raw = relative_expression
    relative_expression = relative_expression_mean_for_biological_replicates(relative_expression)
       
 ###Calculate change in expression according to a set comparison value   
compare_basename = 'PS'
gene_names = set()
for outer_key, inner_dict in relative_expression.items():
   for inner_key  in inner_dict:
       gene_names.add(inner_key)

#change_in_expression_dict = {}
#for compare_gene_name in gene_names:
    #change_in_expression_dict = change_in_expression(relative_expression, compare_basename, compare_gene_name)


###output in excel
df_result_dict_CT = pd.DataFrame(result_dict_CT)
df_result_dict_CT = df_result_dict_CT.explode(list(df_result_dict_CT.columns))

df_means = pd.DataFrame(means).round(3)
df_relative_expression = pd.DataFrame(relative_expression).round(10)
#df_change_in_expression = pd.DataFrame(change_in_expression_dict).round(3)

re_replicates = relative_expression_per_replicate(result_dict_CT)


tidy_data = []

for strain, gene_data in re_replicates.items():
    for gene, re_values in gene_data.items():
        for i, value in enumerate(re_values):
            tidy_data.append({
                "Strain": strain,
                "Gene": gene,
                "Replicate": i + 1,
                "RelativeExpression": value
            })

df_tidy = pd.DataFrame(tidy_data)


base_name = os.path.splitext(os.path.basename(input_file_path))[0]  # Get the base name without extension
output_file_path_full = os.path.join(output_file_path, f"{base_name}_RTq_PCR_analysis.xlsx")    
    
    

# Write headers as text at specific rows before the DataFrames
#wb = Workbook()
#sheet = wb.active
#sheet.title = "Sheet1"
#wb.save(output_file_path_full)

with pd.ExcelWriter(output_file_path_full, engine='openpyxl', mode='w') as writer:  
    df_result_dict_CT.to_excel(writer, sheet_name='Sheet1', startrow=4, startcol=1, index=True)    
    df_means.to_excel(writer, sheet_name='Sheet1', startrow=14, startcol=1, index=True)       
    df_relative_expression.to_excel(writer, sheet_name='Sheet1', startrow=20, startcol=1, index=True)       
    #df_change_in_expression.to_excel(writer, sheet_name='Sheet1', startrow=25, startcol=1, index=True)    
    
wb = load_workbook(output_file_path_full)
ws = wb['Sheet1']    

# Adding header for the first dictionary (CT Values)
ws.cell(row=3, column=3).value = "CT Values for All Replicates"

# Adding header for the second dictionary (Mean CT Values)
ws.cell(row=13, column=3).value = "Mean CT Values for Technical Replicates"

# Adding header for the third dictionary (Relative Expression)
ws.cell(row=19, column=3).value = "Relative Expression for Technical Replicates"

# Adding header for the fourth dictionary (Change in Expression)
#ws.cell(row=24, column=3).value = "Change in Expression"

# Set the column widths based on the max length of content in each column (including the index)
# Handle the index column separately
max_length_index = 0
for cell in ws['A']:  # 'A' is the default column for index in pandas
    try:
        if cell.value:
            max_length_index = max(max_length_index, len(str(cell.value)))
    except:
        pass
adjusted_width_index = (max_length_index + 2)  # Add some padding
ws.column_dimensions['B'].width = adjusted_width_index

# Set the column widths based on the max length of content in each column
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name (A, B, C, etc.)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)  # Add some padding to the width
    ws.column_dimensions[column].width = adjusted_width


# Save the workbook
wb.save(output_file_path_full)    

###plot relative expression    
#------------------------------------------------------------------------------
#df_relative_expression['CC-1883'] = 0
df_relative_expression = df_relative_expression.drop(columns=['water'])
#df_error_bars['CC-1883'] = 0

plot_data = df_relative_expression.iloc[0] 
plot_data = plot_data.fillna(0)
#plot_data = plot_data.drop('water') 
error_data = df_error_bars.iloc[0] 
error_data = error_data.drop('water') 

# Define the Excel output path
output_excel_path = os.path.join(output_file_path, f"{base_name}_RTq_PCR_plot_data_rE_20-11-24.xlsx")

# Convert plot_data and error_data to DataFrame if they are Series
plot_data_df = plot_data.to_frame(name='Relative Expression') if isinstance(plot_data, pd.Series) else plot_data
error_data_df = error_data.to_frame(name='Error') if isinstance(error_data, pd.Series) else error_data

# Save both to separate sheets
with pd.ExcelWriter(output_excel_path, engine='xlsxwriter') as writer:
    plot_data_df.to_excel(writer, sheet_name='Plot_Data')
    error_data_df.to_excel(writer, sheet_name='Error_Bars')
    df_tidy.to_excel(writer, sheet_name='replicates')


# Define a custom legend entry
#legend_patch = Patch(color='#B200FF', label='t1 = 59 days')
    
plt.figure(figsize=(10, 7))
x_positions = np.arange(len(df_relative_expression.columns)) 
# Cleaned x labels (removing trailing underscores)
x_labels = df_relative_expression.columns.str.replace(r'_$', '', regex=True)

plt.ylim(0, 0.04)

colors = ['#B200FF' if i < 4 else '#B200FF' for i in range(len(df_relative_expression.columns))]

bars = plt.bar(x_labels,
    plot_data,
    yerr=error_data,
    capsize=10,
    color=colors,
    #edgecolor='black',
    linewidth=1)  # mint color

# Add labels and title
plt.xlabel('Strain', fontsize=24, labelpad=20)
plt.ylabel('rE (compared to RPL13)', fontsize=24, labelpad=20)
#plt.title('Relative expression of codon_opt_L9 of LHCBM9 OEx and LHCBM9+YFP OEx mutants',
          #fontsize=14)
plt.xticks(rotation = 45, fontsize=18, ha='right')
plt.yticks(fontsize=18)

# Add a vertical dotted line after the fourth bar (index 3, between 3 and 4)
plt.axvline(x=3.5, color='#333333', linestyle='dotted', linewidth=2)
plt.axvline(x=8.5, color='#333333', linestyle='dotted', linewidth=2)


previous_height = None  # Use None to handle the first iteration

for bar in bars:
    height = bar.get_height()
    offset = 0.003  # Default offset

    if previous_height is not None:
        diff = height - previous_height
        if height == 0.0034975111:
            offset = 0.002   
        elif abs(diff) < 0.0018:
            if diff > 0:
                offset = 0.005
            else:
                offset = 0.002
                
                
        
    print(offset)
    
    plt.text(
        bar.get_x() + bar.get_width() / 2,  # X-coordinate: center of the bar
        height + offset,                            # Y-coordinate: top of the bar
        f"{height:,.4f}",                # Format the value with commas
        ha='center', va='bottom',          # Align text to be centered
        fontsize=18, color='black'
    )
    
    previous_height = height

# Add the legend to the plot
plt.legend(
    #handles=[legend_patch],
    title="Gene",
    fontsize=18,
    loc='upper left',
    frameon=True,
    fancybox=True,
    title_fontsize=20,
)

# Save the plot
plt.tight_layout()
output_file_path_pic = os.path.join(output_file_path, f"{base_name}_RTq_PCR_analysis_plot.png")    
plt.savefig(output_file_path_pic, dpi=300, format='png')  # Save as PNG with high resolution

# Show the plot
plt.show()    
    
###YFP plot only
#------------------------------------------------------------------------------
# Extract data specific to YFP
yfp_data = df_relative_expression.filter(like='YFP')  # Filter columns containing 'YFP'
plot_data2 = yfp_data.iloc[0]  # Assuming the first row contains the data to plot

# Plot
plt.figure(figsize=(10, 6))
x_positions = np.arange(len(yfp_data.columns))  # X positions for YFP data
colors = ['#FFC300'] * len(yfp_data.columns)   # Mint color for all bars
bars = plt.bar(yfp_data.columns.str.replace(r'_$', '', regex=True), plot_data2, color=colors)

# Add labels and title
plt.xlabel('Strain', fontsize=12, labelpad=20)
plt.ylabel('rE (compared to RPL13)', fontsize=12, labelpad=20)
plt.title('rE (compared to RPL13) of codon_opt_L9 for LHCBM9+YFP constructs', fontsize=14)
plt.xticks(rotation=45, fontsize=12)
plt.yticks(fontsize=12)

# Format the y-axis using scientific notation
#ax = plt.gca()
#formatter = ScalarFormatter(useMathText=True)  # Use scientific notation
#formatter.set_powerlimits((-6, 6))             # Set limits for scientific notation
#ax.yaxis.set_major_formatter(formatter)
#ax.ticklabel_format(axis='y', style='sci', scilimits=(6, 6))  # Use exponential notation for 1e6

# Add values on top of the bars
for bar in bars:
    height = bar.get_height()
    plt.text(
        bar.get_x() + bar.get_width() / 2,
        height,
        f"{height:,.4f}",
        ha='center', va='bottom', fontsize=10, color='black'
    )

# Save the plot
plt.tight_layout()
yfp_output_file_path = os.path.join(output_file_path, f"{base_name}_RTq_PCR_YFP_plot.png")
plt.savefig(yfp_output_file_path, dpi=300, format='png')

# Show the plot
plt.show()