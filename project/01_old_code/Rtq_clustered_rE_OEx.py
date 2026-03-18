#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Apr 13 16:11:48 2025

@author: wweber
"""

import pandas as pd
import numpy as np
import os
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
import re
from datetime import datetime
from openpyxl.drawing.image import Image
from collections import defaultdict
from itertools import cycle
from decimal import Decimal
from collections import Counter
from pathlib import Path
import matplotlib
import matplotlib.ticker as ticker
from matplotlib.lines import Line2D
import glob
from scipy import stats
import statsmodels.api as sm
from statsmodels.formula.api import ols
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from scipy.stats import kruskal
from scipy.stats import shapiro, levene
import scikit_posthocs as sp
from statsmodels.stats.multitest import multipletests
from itertools import combinations
import matplotlib.patches as patches
import matplotlib.colors as mcolors
from scipy.stats import ttest_ind

from scipy.stats import levene, bartlett
import numpy as np
from scipy import stats
from scipy.stats import shapiro, levene, f_oneway, kruskal
import pingouin as pg
import scikit_posthocs as sp
from statsmodels.stats.multicomp import pairwise_tukeyhsd
from openpyxl.utils import get_column_letter


def check_stats_and_test_with_posthoc(
df,
    group_col,
    value_col,
    stratify_by=None,
    alpha=0.05
):
    summary_results = []
    posthoc_results = []
    normality_detail_records = []

    # Convert to numeric and drop NaNs
    df[value_col] = pd.to_numeric(df[value_col], errors='coerce')
    df = df.dropna(subset=[value_col])

    # Define strata
    strata_values = df[stratify_by].unique() if stratify_by else [None]
    num_tests = len(strata_values)

    for stratum in strata_values:
        sub_df = df[df[stratify_by] == stratum] if stratify_by else df.copy()
        stratum_label = stratum if stratify_by else "All Data"
        unique_groups = sub_df[group_col].unique()

        # Normality per group
        normality_flags = []
        for grp in unique_groups:
            grp_data = sub_df[sub_df[group_col] == grp][value_col]
            if len(grp_data) >= 3:
                norm_res = pg.normality(grp_data)
                normality_flags.append(norm_res["normal"].values[0])

                normality_detail_records.append({
                    "Comparison": stratum_label,
                    "Group": grp,
                    "W-statistic": norm_res["W"].values[0],
                    "p-value": norm_res["pval"].values[0],
                    "Normal": norm_res["normal"].values[0]
                })
            else:
                normality_flags.append(False)
                normality_detail_records.append({
                    "Comparison": stratum_label,
                    "Group": grp,
                    "W-statistic": None,
                    "p-value": None,
                    "Normal": False
                })

        is_normal = all(normality_flags)
        
        if len(unique_groups) == 2:
            group1_name, group2_name = unique_groups
            group1 = sub_df[sub_df[group_col] == group1_name][value_col]
            group2 = sub_df[sub_df[group_col] == group2_name][value_col]
        
            if is_normal:
                levene_res = pg.homoscedasticity(sub_df, dv=value_col, group=group_col)
                homogeneity = levene_res["equal_var"].values[0]
        
                if homogeneity:
                    stat, pval_overall = stats.ttest_ind(group1, group2)
                    test_type = "t-test"
                else:
                    stat, pval_overall = stats.ttest_ind(group1, group2, equal_var=False)
                    test_type = "Welch t-test"
            else:
                stat, pval_overall = stats.mannwhitneyu(group1, group2, alternative='two-sided')
                test_type = "Mann–Whitney U"
                homogeneity = False
                
            p_adj = min(pval_overall * num_tests, 1.0)
            
            # Record overall summary
            summary_results.append({
                "Comparison": stratum_label,
                "Normal": is_normal,
                "Homogeneous": homogeneity,
                "Test": test_type,
                "p-value": p_adj,
                "Significant": pval_overall < alpha,
            })
        
            # Add posthoc-like result to match multi-group structure
            posthoc_results.append(pd.DataFrame([{
                "group1": group1_name,
                "group2": group2_name,
                "p-unc": pval_overall,
                "p-adj": p_adj,  # No adjustment needed for 1 test
                "Test": test_type,
                "Comparison": stratum_label,
                "Significant": pval_overall < alpha,
                'Bonf_sig': p_adj < alpha
            }]))
        
            continue  # Skip multi-group logic    
        
        # Homogeneity of variance
        if len(unique_groups) > 2:
            levene = pg.homoscedasticity(sub_df, dv=value_col, group=group_col)
            homogeneity = levene["equal_var"].values[0]
        else:
            homogeneity = True
        
        # Choose and apply statistical test
        if is_normal and homogeneity:
            test_type = "ANOVA"
            stat_res = pg.anova(dv=value_col, between=group_col, data=sub_df)
            pval_overall = stat_res["p-unc"].values[0]

            if pval_overall < alpha:
                # Posthoc: Tukey HSD
                tukey = pairwise_tukeyhsd(sub_df[value_col], sub_df[group_col])
                tukey_df = pd.DataFrame(
                    data=tukey._results_table.data[1:],
                    columns=tukey._results_table.data[0]
                )
                # rename for clarity
                tukey_df = tukey_df.rename(columns={'p-adj': 'p-adj'})
                tukey_df['p-unc'] = np.nan  # Tukey only provides adjusted
                tukey_df["Test"] = "Tukey HSD"
                tukey_df["Comparison"] = stratum_label
                tukey_df["Significant"] = tukey_df["p-adj"] < alpha
                posthoc_results.append(tukey_df)

        elif is_normal and not homogeneity:
            test_type = "Welch ANOVA"
            stat_res = pg.welch_anova(dv=value_col, between=group_col, data=sub_df)
            pval_overall = stat_res["p-unc"].values[0]

            if pval_overall < alpha:
                # Posthoc: Games-Howell
                gh = pg.pairwise_gameshowell(dv=value_col, between=group_col, data=sub_df)
                # rename pingouin's p-unc column
                gh = gh.rename(columns={'pval': 'p-unc'})
                gh['p-adj'] = gh['p-unc']  # Games-Howell by default unadjusted
                gh["Test"] = "Games-Howell"
                gh["Comparison"] = stratum_label
                gh["Significant"] = gh["p-unc"] < alpha
                posthoc_results.append(gh)

        else:
            test_type = "Kruskal-Wallis"
            stat_res = pg.kruskal(dv=value_col, between=group_col, data=sub_df)
            pval_overall = stat_res["p-unc"].values[0]

            if pval_overall < alpha:
                # Posthoc: Dunn test (Bonferroni)
                dunn_adj = sp.posthoc_dunn(
                    sub_df, val_col=value_col, group_col=group_col, p_adjust="bonferroni"
                )
                dunn = dunn_adj.stack().reset_index()
                dunn.columns = ["group1", "group2", "p-adj"]
                # now run raw (no adjustment) for unadjusted p-values
                dunn_raw = sp.posthoc_dunn(
                    sub_df, val_col=value_col, group_col=group_col, p_adjust=None
                ).stack().reset_index()
                dunn_raw.columns = ["group1", "group2", "p-unc"]

                # merge raw+adj, drop self-comparisons
                merged = pd.merge(dunn_raw, dunn, on=["group1", "group2"])
                merged = merged[merged["group1"] != merged["group2"]]
                merged["Test"] = "Dunn"
                merged["Comparison"] = stratum_label
                merged["Significant"] = merged["p-adj"] < alpha
                posthoc_results.append(merged)

        # Add overall summary
        summary_results.append({
            "Comparison": stratum_label,
            "Normal": is_normal,
            "Homogeneous": homogeneity,
            "Test": test_type,
            "p-value": pval_overall,
            "Significant": pval_overall < alpha
        })

    # Final outputs
    summary_df = pd.DataFrame(summary_results)
    posthoc_df = pd.concat(posthoc_results, ignore_index=True) if posthoc_results else pd.DataFrame()
    normality_detail_df = pd.DataFrame(normality_detail_records)

    # Ensure grouping columns are strings
    for col in ['group1', 'group2', 'Comparison']:
        if col in posthoc_df.columns:
            posthoc_df[col] = posthoc_df[col].astype(str)

    return summary_df, posthoc_df, normality_detail_df

def export_to_excel_with_autofit(filename, df_dict):
    """
    Export multiple DataFrames to an Excel file, each in its own sheet,
    with columns width auto-adjusted based on the longest text in each column.

    Parameters:
    - filename: str, path to the output Excel file (e.g., 'results.xlsx')
    - df_dict: dict, keys = sheet names, values = DataFrames to write
    """
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in df_dict.items():
            for col in df.select_dtypes(include='bool').columns:
                df[col] = df[col].map({True: "True", False: "False"})
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]

            for i, col in enumerate(df.columns, 1):
                # Find max length of the column header and all column values (converted to str)
                max_len = max(
                    df[col].astype(str).map(len).max(),
                    len(str(col))
                )
                # Add a little extra space (2 chars)
                adjusted_width = max_len + 4
                # Set column width (Excel columns are 1-indexed)
                col_letter = get_column_letter(i)
                worksheet.column_dimensions[col_letter].width = adjusted_width    

file_paths = [
    '/prj/lhcbm9/02_Master_project_08-2024_03-2025/04_RTq_PCR/results_20-11-24_RTq_PCR_plot_data_rE_20-11-24.xlsx',
    '/prj/lhcbm9/02_Master_project_08-2024_03-2025/04_RTq_PCR/01_OEx_vs_PS-S/-S_comparsison_raw_data_RTq_PCR_plot_data_rE_07-01-25.xlsx'
]


#---------------------------------------------------------------------------------------------------
# Initialize a list to collect all data
combined_data = []

# Read and parse each file
for path in file_paths:
    # Extract date from filename (e.g., 2024-04-10)
    match = re.search(r"\d{2}-\d{2}-\d{2}", os.path.basename(path))
    if match:
        file_date = match.group(0)
    else:
        continue

    # Read both sheets
    df_expression = pd.read_excel(path, sheet_name='Plot_Data')
    df_errors = pd.read_excel(path, sheet_name='Error_Bars')
    
    # Clean up column names
    df_expression.rename(columns={"Unnamed: 0": "Basename"}, inplace=True)
    df_errors.rename(columns={"Unnamed: 0": "Basename"}, inplace=True)

    # Add a column to tag with date
    df_expression["Date"] = file_date
    df_errors["Date"] = file_date
    

    # Merge expression and error data on common columns (assume 'BaseName' is common)
    df_merged = pd.merge(df_expression, df_errors, on=["Basename", "Date"], suffixes=("", "_error"))

    # Add to list
    combined_data.append(df_merged)

# Combine all into one big DataFrame
df_all = pd.concat(combined_data, ignore_index=True)

rename_dict = {'pML2_603_' : 'YFP+L9-OEx-',
               'pML2_604_' : 'L9-OEx-',
               'CC-1883' : 'PS'
    }

for old, new in rename_dict.items():
    df_all['Basename'] = df_all['Basename'].str.replace(old, new, regex=True)




#---------------------------------------------------------------------------------------------------
# To collect all replicate values for statistical testing
replicate_data = []

for path in file_paths:
    match = re.search(r"\d{2}-\d{2}-\d{2}", os.path.basename(path))
    if match:
        file_date = match.group(0)
    else:
        continue

    # --- Sheet 3: Replicates ---
    df_reps = pd.read_excel(path, sheet_name=2)  # Sheet 3 is index 2

    # Melt wide to long format
    df_reps = df_reps.drop(columns=["Unnamed: 0"])
    df_reps["Date"] = file_date

    replicate_data.append(df_reps)

# Combine all replicate info
df_replicates_all = pd.concat(replicate_data, ignore_index=True)
df_replicates_all["Gene"] = df_replicates_all["Gene"].replace("LHCBM9", "codon_opt_L9")
for old, new in rename_dict.items():
    df_replicates_all['Strain'] = df_replicates_all['Strain'].str.replace(old, new, regex=True)
df_codon_opt = df_replicates_all[df_replicates_all["Gene"] == "codon_opt_L9"]

# Statistics
#---------------------------------------------------------------------------------------------------
summary_df, posthoc_df, normality_detail_df = check_stats_and_test_with_posthoc(
    df=df_codon_opt,
    value_col='RelativeExpression',
    group_col= 'Date',
    stratify_by= 'Strain',
    alpha=0.05
)

#---------------------------------------------------------------------------------------------------
# OEx codon opt L9
palette_dict = {'t1 = 59 days' :'#B200FF',
                't2 = 116 days' : '#FF1493'}

# natural L9
palette_dict_2 = {'t1 = 59 days' :'#00BFB6',
                't2 = 116 days' : '#00FFCE'}

# Updated, high-contrast, distinguishable colors
custom_palette_basenames = [
    "#FFBB00",  # yellow CC-1883
    "#80ffdb", "#07BEB8",  # turquoise KOs
    "#80CBFF", "#0A84E6", "#003A7E"   # blue shades OEx
    
]

base_names_list = df_all['Basename'].unique()

date_time_point_dict = {'20-11-24' : 't1 = 59 days',
                   '07-01-25' : 't2 = 116 days'}

df_all['Date'] = df_all['Date'].replace(date_time_point_dict)
#---------------------------------------------------------------------------------------------------
# Plotting clustered barplot

x_pos = []

plt.figure(figsize=(10, 6))

colors = ['#FFC300' if i < 4 else '#2E83CC' for i in range(int(len(df_all)/2))]

plt.ylim(0, 0.04)

ax = sns.barplot(
    data=df_all,
    x="Basename", y="Relative Expression", hue="Date",
    errorbar=None,
    palette= palette_dict)

# Now add error bars on top of actual bars
for bar, (_, row) in zip(ax.patches, df_all.iterrows()):
    x = bar.get_x() + bar.get_width() / 2
    x_pos.append(x)
    y = bar.get_height()
    err = row["Error"]

    ax.errorbar(
        x, y, yerr=err,
        fmt='none', capsize=10, color='black', linewidth=1
    )
    
    ax.errorbar(x, y, yerr=err, fmt='none', capsize=10, color='black', linewidth=1)

    # Pick custom offset based on the 'Date' column
    date_offset = row["Date"]
    offset_y = 0.002 if row["Date"] == 't1 = 59 days' else 0.006
    offset_x = +0.1 if row['Date'] == 't1 = 59 days' else -0.1
    
    plt.text(
        x + offset_x,  # X-coordinate: center of the bar
        y + offset_y,                            # Y-coordinate: top of the bar
        f"{y:,.4f}",                # Format the value with commas
        ha='center', va='bottom',          # Align text to be centered
        fontsize=12, color='black'
    )
    
plt.legend(title="Time after\ntransformation",
    fontsize=14,
    loc='upper left',
    frameon = True,
    fancybox = True,
    #bbox_to_anchor=(0.5, -0.6),
    #columnspacing = 10,
    ncol=1, 
    title_fontsize=16,
    )

x_pos = sorted(x_pos)

x_pos_strain = []

# Iterate through the list in steps of 2
for i in range(0, len(x_pos), 2):
    # Average the current value and the next one (if it exists)
    avg = (x_pos[i] + x_pos[i + 1]) / 2 if i + 1 < len(x_pos) else x_pos[i]
    x_pos_strain.append(avg)


for i, row in posthoc_df.iterrows():
    if row["p-unc"] < 0.05:
        basename = row["Comparison"]
        t = list(base_names_list).index(basename)
        x = x_pos_strain[t]
        x1 =x_pos[t*2]
        x2 = x_pos[t*2+1]
        # Access the Relative Expression for each strain (using .loc[])
        y1 = df_all.loc[df_all["Basename"] == basename].iloc[0]["Relative Expression"]
        y2 = df_all.loc[df_all["Basename"] == basename].iloc[1]["Relative Expression"]
        if y1 < 0.015 and y2 < 0.015:
            y = 0.015
        else:
            y = max(y1, y2) + 0.005
        
        if row["p-adj"] < 0.05:
            # Draw a line between bars
            ax.plot([x1, x1, x2, x2], [y, y+0.001, y+0.001, y], lw=1.5, c='black')
            ax.text(x, y+0.0015, "*", ha='center', va='bottom', color='red', fontsize=22)
            
        else:
            # Draw a line between bars
            ax.plot([x1, x1, x2, x2], [y, y+0.001, y+0.001, y], lw=1.5, c='black')
            ax.text(x, y+0.0015, "*", ha='center', va='bottom', color='black', fontsize=22)
    

# Add a vertical dotted line after the fourth bar (index 3, between 3 and 4)
plt.axvline(x=3.5, color='#333333', linestyle='dotted', linewidth=2)
plt.axvline(x=8.5, color='#333333', linestyle='dotted', linewidth=2)


plt.xlabel('Strain', fontsize=18, labelpad=20)
plt.ylabel('rE (compared to RPL13)', fontsize=18, labelpad=20)

plt.xticks(rotation = 45, fontsize=14, ha = 'right')
plt.yticks(fontsize=14)

# Save the plot
plt.tight_layout()
output_file_path_pic = "/prj/lhcbm9/02_Master_project_08-2024_03-2025/14_thesis_plots/RTq_PCR_analysis_plot_clustered_rE_OEx_gene_silencing.png"   
plt.savefig(output_file_path_pic, dpi=300, format='png')  # Save as PNG with high resolution

# Show the plot
plt.show()    
